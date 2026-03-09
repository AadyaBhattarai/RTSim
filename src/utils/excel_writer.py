"""
Excel output utilities for simulation results.
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def sanitize_workbook(path: str) -> None:
    """Remove corrupted workbook if it exists."""
    if os.path.exists(path):
        try:
            load_workbook(path)
        except Exception:
            print(f"  Warning: removing corrupted workbook: {path}")
            os.remove(path)


def append_df_to_excel(df: pd.DataFrame, filename: str, sheet_name: str) -> None:
    """
    Append a DataFrame to an Excel file.

    Creates the file and sheet if they don't exist.
    Appends rows (without header) if the sheet already exists.
    """
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(filename)
        return

    wb = load_workbook(filename)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    wb.save(filename)
