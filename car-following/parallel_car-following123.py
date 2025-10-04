import re
import subprocess
import traci
import sumolib.miscutils as mu
import pandas as pd
import numpy as np
import random
from scipy import stats
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils7 import add_platooning_vehicle, start_sumo, running, communicate, add_vehicle
from plexe import Plexe, ACC, CACC, PLOEG, CONSENSUS, RPM, GEAR, RADAR_DISTANCE, ENGINE_MODEL_REALISTIC
import secrets
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
import os
 
# Debug logging
debug_logs = []
def debug_print(msg):
    print(msg)
    debug_logs.append(msg)
 
BASE_PHEM_DIR = r"/home/a/aadyab/ab/sumo_source/data/emissions"
CRR_VALUES = {'primary': 0.006923, 'secondary': 0.010, 'cross_country': 0.025}
ROAD_TYPES = ['primary']  # Only one road type at a time
 
output_dir = "experiment1_car_following_results123_prim"
os.makedirs(output_dir, exist_ok=True)
RAW_FILE = os.path.join(output_dir, "all_scenarios_raw.xlsx")
 
def ci_file_for_model(model):
    return os.path.join(output_dir, f"confidence_intervals_{model}.xlsx")
 
def sanitize_workbook(path):
    """If path exists but is not a valid .xlsx, delete it so we can recreate cleanly."""
    if os.path.exists(path):
        try:
            load_workbook(path)
        except Exception:
            print(f"⚠️  Removing corrupted workbook: {path}")
            os.remove(path)
 
# Sanitize at startup
sanitize_workbook(RAW_FILE)
for model in ["Model1", "Model2", "Model3","Model4", "Model5"]:
    sanitize_workbook(ci_file_for_model(model))
 
def append_df_to_excel(df, filename, sheet_name):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(filename)
        return
 
    wb = load_workbook(filename)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    wb.save(filename)
 
def generate_sumo_network(input_file, base_output, slopes):
    nets = []
    for s in slopes:
        out = f"{base_output}_slope_{s}.net.xml"
        data = open(input_file).read()
        data = re.sub(r'b=".*?"', f'b="{s}"', data)
        tmp = "temp.xodr"
        open(tmp, 'w').write(data)
        subprocess.run(
            f"netconvert --opendrive-files {tmp} --ignore-errors -o {out}",
            shell=True, check=True
        )
        os.remove(tmp)
        nets.append((s, out))
    return nets
 
def generate_new_config(cfg_file, route_file, net_file):
    base_cfg = open(cfg_file).read()
    rv = route_file.replace('\\', '/')
    nv = net_file.replace('\\', '/')
    updated = re.sub(
        r'<route-files\s+value="[^"]+"',
        f'<route-files value="{rv}"',
        base_cfg
    )
    updated = re.sub(
        r'<net-file\s+value="[^"]+"',
        f'<net-file value="{nv}"',
        updated
    )
    unique_name = f"sumo_{secrets.token_hex(8)}.sumocfg"
    with open(unique_name, 'w', encoding='utf-8') as f:
        f.write(updated)
        f.flush()
        os.fsync(f.fileno())
    return unique_name
 
 
def generate_car_following_route_files(models, truck_counts, sigma_tau_pairs, accels, speeds, min_gaps):
    base =  r"/home/a/aadyab/ab/carfollowing"
    output_dir = "generated_car_following_routes123"
    os.makedirs(output_dir, exist_ok=True)
    files = []
    for m in models:
        for c in truck_counts:
            for v in ['lower', 'upper']:
                if c == 1:
                    tpl = os.path.join(base, m, f"{c}truck", v, "grade2.rou.xml")
                    if not os.path.exists(tpl):
                        continue
                    txt = open(tpl).read()
                    for sigma, tau in sigma_tau_pairs:
                        for accel in accels:
                            for speed in speeds:
                                # txt = open(tpl).read()
                                speed_ms = speed / 3.6
                                txt = re.sub(r'sigma="[^"]+"', f'sigma="{sigma}"', txt)
                                txt = re.sub(r'tau="[^"]+"', f'tau="{tau}"', txt)
                                txt = re.sub(r'accel="[^"]+"', f'accel="{accel}"', txt)
                                txt = re.sub(r'maxSpeed="[^"]+"', f'maxSpeed="{speed_ms}"', txt)
                                txt = re.sub(r'departSpeed="[^"]+"', f'departSpeed="{speed_ms}"', txt)
                                name = (
                                    f"route_{m}_{c}truck_{v}"
                                    f"_sigma_{sigma}"
                                    f"_tau_{tau}"
                                    f"_accel_{accel}"
                                    f"_maxSpeed_{speed}"
                                    f"_minGap_0.rou.xml"
                                )
                                out = os.path.join(output_dir, name)
                                open(out, 'w').write(txt)
                                files.append(out)
                else:
                    for g in min_gaps:
                        tpl = os.path.join(base, m, f"{c}truck", v, "grade2.rou.xml")
                        if not os.path.exists(tpl):
                            continue
                        txt = open(tpl).read()
                        for sigma, tau in sigma_tau_pairs:
                            for accel in accels:
                                for speed in speeds:
                                    # txt = open(tpl).read()
                                    speed_ms = speed / 3.6
                                    txt = re.sub(r'sigma="[^"]+"', f'sigma="{sigma}"', txt)
                                    txt = re.sub(r'tau="[^"]+"', f'tau="{tau}"', txt)
                                    txt = re.sub(r'accel="[^"]+"', f'accel="{accel}"', txt)
                                    txt = re.sub(r'maxSpeed="[^"]+"', f'maxSpeed="{speed_ms}"', txt)
                                    txt = re.sub(r'departSpeed="[^"]+"', f'departSpeed="{speed_ms}"', txt)
                                    txt = re.sub(r'minGap="[^"]+"', f'minGap="{g}"', txt)
                                    name = (
                                        f"route_{m}_{c}truck_{v}"
                                        f"_sigma_{sigma}"
                                        f"_tau_{tau}"
                                        f"_accel_{accel}"
                                        f"_maxSpeed_{speed}"
                                        f"_minGap_{g}.rou.xml"
                                    )
                                    out = os.path.join(output_dir, name)
                                    open(out, 'w').write(txt)
                                    files.append(out)
    return files
 
def calculate_confidence_interval(data, confidence=0.95):
    arr = np.array(data, dtype=float)
    if len(arr) < 2 or np.all(np.isnan(arr)):
        return np.nan, np.nan, np.nan, len(arr)
    mean = np.nanmean(arr)
    std = np.nanstd(arr, ddof=1)
    n = np.sum(~np.isnan(arr))
    sem = std / np.sqrt(n)
    t_crit = stats.t.ppf((1 + confidence) / 2, df=n-1)
    return mean, mean - t_crit*sem, mean + t_crit*sem, n
 
def run_car_following_simulation(cfg, route, seed):
    pat = re.compile(
        r'^route_(\w+)_(\d+)truck_(lower|upper)_sigma_([\d\.]+)_tau_([\d\.]+)_accel_([\d\.]+)_maxSpeed_([\d\.]+)_minGap_([\d\.]+)\.rou\.xml$'
    )
    m = pat.match(os.path.basename(route))
    if not m:
        raise ValueError(f"Bad route: {route}")
    # Extract parameters from the route filename
    model, cnt_str, variant, sigma_str, tau_str, accel_str, speed_str, gap_str = m.groups()
    cnt, sigma, tau, accel, speed = int(cnt_str), float(sigma_str), float(tau_str), float(accel_str), int(speed_str)
    gap = float(gap_str) if gap_str else 0.0
 
    sumo_cmd = ["sumo", "-c", cfg]
    traci.start(sumo_cmd)
    cum_fuel, cum_dist = {}, {}
    while traci.simulation.getMinExpectedNumber() > 0:
        traci.simulationStep()
        for vid in traci.vehicle.getIDList():
            cum_fuel[vid] = cum_fuel.get(vid, 0.0) + traci.vehicle.getFuelConsumption(vid)*traci.simulation.getDeltaT()
            cum_dist[vid] = traci.vehicle.getDistance(vid)/1000.0
    traci.close()
 
    results = []
    for vid in cum_fuel:
        km = cum_dist.get(vid, 0.0)
        fuel_L100km = (cum_fuel[vid]/850000.0/km*100) if km>0 else None
        results.append({
            'Scenario': 'car_following', 'Model': model, 'Count': cnt,
            'Sigma': sigma, 'Tau': tau, 'Accel': accel, 'Speed': speed,
            'minGap': gap, 
            'Vehicle': vid,
            'Fuel_L_per_100km': fuel_L100km, 'Distance_km': km
        })
    return results
 
def simulate_one(slope, net, road, rt, trial, base_cfg):
    cfg_path = None
    try:
        cfg_path = generate_new_config(base_cfg, rt, net)
        rows = run_car_following_simulation(cfg_path,rt,trial)
        for r in rows:
            r.update({'Slope': slope, 'RoadType': road})
        print(f"[Trial {trial}] done: slope={slope}, road={road}, route={os.path.basename(rt)}")
        return rows
    except Exception as e:
        debug_print(f"[ERROR] Trial {trial} failed: {e}")
        return []
    finally:
        if cfg_path and os.path.exists(cfg_path):
            try:
                os.remove(cfg_path)
            except Exception as e:
                debug_print(f"Failed to delete temporary config file {cfg_path}: {e}")
 
if __name__ == "__main__":
    from tqdm import tqdm
 
    models = ["Model1", "Model2", "Model3"]
    truck_counts = [3]
    sigma_tau_pairs = [
        (0.5, 1.00), (0.4, 0.95), (0.3, 0.90),
        (0.2, 0.85), (0.0, 0.80), (0.0, 0.75),
    ]
    accel_values = [0.2, 0.6, 1.0]
    speeds_km_h = [30, 60, 90, 100]
    minGap_values = [5, 10, 15, 20]
    slope_values = [0.0, 0.06, 0.08, 0.16, 0.20]
    base_cfg = "grade0.sumocfg"
    input_xodr = "experiment_e.xodr"
    base_net_str = "simulation_network"
    trials = 1
 
    nets = generate_sumo_network(input_xodr, base_net_str, slope_values)
    platoon_rts = generate_car_following_route_files(models, truck_counts, sigma_tau_pairs, accel_values, speeds_km_h, minGap_values)
 
    scenario_param_list = []
    for slope_val, net in nets:
        for road in ROAD_TYPES:
            for sp in speeds_km_h:
                for g in minGap_values:
                    for sigma, tau in sigma_tau_pairs:
                        for accel in accel_values:
                            scenario_param_list.append({
                                "slope": slope_val, "net": net, "road": road, "speed": sp, 
                                "minGap": g,
                                  "sigma": sigma, "tau": tau, "accel": accel
                            })
 
    total_scenarios = len(scenario_param_list)
    model_bars = {model: tqdm(total=total_scenarios, desc=f"{model} scenarios", position=i+1, leave=False)
                  for i, model in enumerate(models)}
    scenario_bar = tqdm(total=total_scenarios, desc="Scenarios completed", position=0)
 
    for scenario_params in scenario_param_list:
        # Collect all route files for this scenario across all variants and models
        route_files = []
        for variant in ['lower', 'upper']:
            for model in models:
                rt = None
                for candidate in platoon_rts:
                    name = os.path.basename(candidate)
                    if (f"route_{model}_" in name and
                        f"sigma_{scenario_params['sigma']}" in name and
                        f"tau_{scenario_params['tau']}" in name and
                        f"accel_{scenario_params['accel']}" in name and
                        f"maxSpeed_{scenario_params['speed']}" in name and
                        f"minGap_{scenario_params['minGap']}" in name and
                        f"_{variant}_" in name):
                        rt = candidate
                        print(f"route file found: {rt}")
                        break
                if rt:
                    route_files.append(rt)
                    print(f"appended route file: {rt}")
 
        # Extract unique emission classes from these route files
        emission_classes = set()
        for rt in route_files:
            with open(rt, 'r') as f:
                content = f.read()
            classes = re.findall(r'emissionClass="([^\"]+)"', content)
            emission_classes.update(classes)
 
        # Modify the .veh files for these emission classes
        crr = CRR_VALUES.get(scenario_params['road'])
        if crr is not None:
            for cls in emission_classes:
                veh_path = os.path.join(BASE_PHEM_DIR, cls + ".PHEMLight.veh")
                if not os.path.isfile(veh_path):
                    print(f"NOT FOUND: {veh_path}")
                    continue
                with open(veh_path, 'r') as vf:
                    lines = vf.read().splitlines()
                new_lines = []
                skip = False
                for line in lines:
                    if skip:
                        skip = False
                        continue
                    if line.strip().lower() == 'c fr0':
                        new_lines.append(line)
                        new_lines.append(str(crr))
                        skip = True
                    else:
                        new_lines.append(line)
                with open(veh_path, 'w') as vf:
                    vf.write("\n".join(new_lines))
                print(f"Modified CRR in: {veh_path} to {crr}")
 
        # Run all simulations for this scenario
        model_results = {m: [] for m in models}
        for variant in ['lower', 'upper']:
            jobs = []
            for model in models:
                rt = None
                for candidate in platoon_rts:
                    name = os.path.basename(candidate)
                    if (f"route_{model}_" in name and
                        f"sigma_{scenario_params['sigma']}" in name and
                        f"tau_{scenario_params['tau']}" in name and
                        f"accel_{scenario_params['accel']}" in name and
                        f"maxSpeed_{scenario_params['speed']}" in name and
                        f"minGap_{scenario_params['minGap']}" in name and
                        f"_{variant}_" in name):
                        rt = candidate
                        break
                if not rt:
                    print(f"Could not find route file for {model}, {scenario_params}, {variant}")
                    continue
                for trial in range(1, trials + 1):
                    jobs.append((model, scenario_params['slope'], scenario_params['net'],
                                 scenario_params['road'], rt,
                                 trial, base_cfg))
 
            if not jobs:
                print(f"No jobs found for scenario {scenario_params}, variant {variant}")
                continue
 
            with ProcessPoolExecutor(max_workers=50) as executor:
                futures = []
                for job in jobs:
                    model, slope, net, road, rt, trial, base_cfg = job
                    fut = executor.submit(simulate_one, slope, net, road, rt, trial, base_cfg)
                    futures.append((fut, model))
                for fut, model in futures:
                    rows = fut.result()
                    if rows:
                        for r in rows:
                            r["Model"] = model
                        model_results[model].extend(rows)
 
        # Calculate confidence intervals
        for model in models:
            all_rows = model_results[model]
            if not all_rows or len(all_rows) < 20:
                print(f"Warning: fewer than 20 rows for {model} in scenario {scenario_params}")
                continue
            df_scn = pd.DataFrame(all_rows)
            ci_records = []
            for veh_id, group in df_scn.groupby('Vehicle'):
                mean = np.nanmean(group['Fuel_L_per_100km'])
                std = np.nanstd(group['Fuel_L_per_100km'], ddof=1)
                n = np.sum(~np.isnan(group['Fuel_L_per_100km']))
                if n > 1:
                    sem = std / np.sqrt(n)
                    t_crit = stats.t.ppf((1 + 0.95) / 2, df=n-1)
                    ci_lower = mean - t_crit * sem
                    ci_upper = mean + t_crit * sem
                else:
                    ci_lower = ci_upper = np.nan
                rec = {
                    **scenario_params,
                    'Model': model,
                    'Vehicle': veh_id,
                    'Mean': mean,
                    'CI_Lower': ci_lower,
                    'CI_Upper': ci_upper,
                    'standard deviation': std,
                    'standard error': sem,
                    'N': n
                }
                ci_records.append(rec)
            if ci_records:
                df_ci = pd.DataFrame(ci_records)
                append_df_to_excel(df_ci, ci_file_for_model(model), sheet_name="CI")
            model_bars[model].update(1)
        scenario_bar.update(1)